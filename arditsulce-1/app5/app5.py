import streamlit as st
from fpdf import FPDF
import pandas as pd
from openpyxl import Workbook

#wb = Workbook("10001-2023.1.18.xlsx")

#ws = wb.

pdf = FPDF(orientation="P", unit="mm", format="Legal")


st.title("Test Page")

df = pd.read_excel("invoices/10001-2023.1.18.xlsx",sheet_name="Sheet 1")
pdf.add_page()

pdf.set_font(family="Times", size=12)
pdf.set_text_color(0, 0, 254)
for index, row in df.iterrows():
    for i in range(row.count()):
        #st.write(f"value of i is {i}")
        #st.write(f"value of row is {row[i]}")
        data = row[i]
        #st.write(data)
        l = len(str(data))
        st.write(f"{data} has length of {l}")
        if i == row.count()-1:
            pdf.cell(w=40, h=12, txt=str(data), align="C", border=1, ln=1)
        else:
            pdf.cell(w=40, h=12, txt=str(data), align="C", border=1)
pdf.output("output.pdf")
#st.write(df)